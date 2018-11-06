import toml as tomLang

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# translate config to dictionary
configDict = tomLang.load('config.toml', _dict=dict)

workbookName = configDict.get(filename)
sourceWorkbook = load_workbook('./Spreadsheets' + workbookName)

sheetName = configDict.get(sheet.name)
sourceSheet = sourceWorkbook.get_sheet_by_name(sheetName)

sourceRow = configDict.get(sheet.rows)
sourceCol = configDict.get(sheet.cols)

outputDict = {}

# iterating through the source sheet
for iterRow in sourceRow:
    for iterCol in sourceCol:

        content = sheet.cell(row = iterRow, column = iterCol)
        colLetter = get_column_letter(iterCol)
        
        # default returnvalue is Error
        mapsTo = configDict.get(colLetter, 'NoMap')
        
        # If source doesn't map to anything, ignore it and move on
        if (mapsTo == 'NoMap'):
            continue
        else:
            # All the values in the cols are in sequential ascending order
            # When transforming them to spreadsheets, just pasting them in 
            # successive rows
            if (outputDict.get(mapsTo, 'None') == None):
                # creating a new array in case nothing exists
                outputDict[mapsTo] = [content]
            else:
                # appending to original array otherwise
                outputDict[mapsTo].append(content)

# TODO: transform the dictionary to a spreadsheet

