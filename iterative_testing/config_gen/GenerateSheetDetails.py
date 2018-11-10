import toml
import openpyxl


def writeSheetOptions(fileObject, name, rows, cols):
    commentString = "# Details of the source sheet. Edit in case info is wrong \n"
    name = "sheet.name = {0}\n".format(name)
    rows = "sheet.rows = +{0}\n".format(str(rows))
    cols = "sheet.cols = {0}\n".format(openpyxl.utils.get_column_letter(cols))
    
    fileObject.write("{0}{1}{2}{3}".format(commentString, name, rows, cols))

// TODO:
def writeHeaders (fileObject):

// TODO:
def writeMaps (fileObject):


def main():
    sourceFile = input("Please enter the name of the source Excel document: ")
    sourceSheet = input(
        "Please enter the name of the sheet you want to be mapped: ")
    # numberOfBoxes = input(
    # "Please enter the number of boxes that need to be shipped: ")
    configFileName = input(
        "Please enter the name of the config file to be generated: ")

    workBook = openpyxl.load_workbook(sourceFile, read_only=True)
    workSheet = workBook[sourceSheet]

    workRows = workSheet.max_row
    workColumns = workSheet.max_column

    # create the toml file
    fileName = configFileName + ".toml"
    file = open(fileName, "w+")

    writeSheetOptions(file, sourceSheet, workRows, workColumns)

    file.close()

if __name__ == "__main__":
    main()
