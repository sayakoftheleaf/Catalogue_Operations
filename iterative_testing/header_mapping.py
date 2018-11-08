import toml

import openpyxl

# from openpyxl import Workbook, load_workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

# translate config to dictionary
configDict = toml.load('config.toml')

# output excel destination
dest = "output_book.xlsx"
wb = openpyxl.Workbook()
outputSheet = wb.create_sheet(title="output_sheet")

# get the header items in a dictionary
headers = configDict.get('header')

# placing all the headers
for key, value in headers.items():
  columnIndex = openpyxl.utils.column_index_from_string(str(key))
  outputSheet.cell(row=1, column=columnIndex, value=str(value))

wb.save('test_book.xlsx')
