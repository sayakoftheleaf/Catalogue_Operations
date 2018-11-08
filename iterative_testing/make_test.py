import toml

import openpyxl

# output excel destination
dest = "input_test.xlsx"
wb = openpyxl.Workbook()
outputSheet = wb.create_sheet(title="output_sheet")

outputSheet.cell(row=1, column=1, value="First")
outputSheet.cell(row=1, column=2, value="Second")
outputSheet.cell(row=2, column=2, value="Third")

wb.save('source.xlsx')
