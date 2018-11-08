import toml
import openpyxl

# translate config to dictionary
configDict = toml.load('config.toml')

# getting the worksheet to transform
sourceWorkbook = openpyxl.load_workbook('./Spreadsheets/' + configDict.get('filename'))

# declaring source information variables
sourceConfig = configDict.get('sheet')
sourceRow = 0
sourceCol = 0
sourceSheet = 'Sheet 1' # default Sheet name in Excel

# loading up the values of the source sheet from the config
for key, value in sourceConfig.items():
  if (str(key) == 'rows'):
    sourceRow = value
  elif (str(key) == 'cols'):
    sourceCol = openpyxl.utils.column_index_from_string(str(value))
  elif (str(key) == 'name'):
    sourceSheet = sourceWorkbook.get_sheet_by_name(str(value))

separator = configDict.get('separator')

# output excel destination
dest = "output_book.xlsx"
wb = openpyxl.Workbook()
outputSheet = wb.create_sheet(title="output_sheet")

headers = configDict.get('header')

# placing all the headers
for key, value in headers.items():
  columnIndex = openpyxl.utils.column_index_from_string(str(key))
  outputSheet.cell(row=1, column=columnIndex, value=str(value))

# iterating through the source sheet
for iterRow in range(1, sourceRow + 1):
  outputDict = {} # empty dictionary for every row
  
  for iterCol in range(1, sourceCol + 1):

    content = sourceSheet.cell(row = iterRow, column = iterCol).internal_value
    if (str(content) == "None"):
      content = "N/A"
    colLetter = openpyxl.utils.get_column_letter(iterCol)
        
    # check if the present column maps to something in the config file
    # default returnvalue is NoMap
    mapsTo = configDict.get('mapto')

    if colLetter in mapsTo:
      # if a mapping exists in config, put the value
      # in the output dictionary
      outputCol = mapsTo[colLetter]
      if outputCol in outputDict:
        outputDict[outputCol].append(content)
      else:
        outputDict[outputCol] = [content]

  for key, dictValue in outputDict.items():
    # adjusting to fit the headers
    normalizedRow = iterRow + 2
    columnIndex = openpyxl.utils.column_index_from_string(key)

    # converting the contents of the cell into a string
    # separated by the specified separator
    if (len(dictValue) > 1):
      formattedValue = separator.join(map(str, dictValue))
    else:
      formattedValue = str(dictValue[0])
    
    # inserting cells into the new
    outputSheet.cell(row = normalizedRow, column = columnIndex, value = formattedValue)

wb.save(dest)
