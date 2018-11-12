import toml
import openpyxl
from pathlib import Path

from generate_box_config import generateExcelColumns
from generate_headers import writeHeaders
from generate_sheet_options import writeSheetOptions
from generate_maps import writeMaps

def main():

    #gathering inputs
    sourceFile = input("Please enter the name of the source Excel document: ")
    
    sourceSheet = input(
        "Please enter the name of the sheet you want to be mapped: ")
    
    numberOfBoxes = input(
        "Please enter the number of boxes that need to be shipped: ")
    
    numberOfBoxes = int(numberOfBoxes)

    sourceBoxesStartFrom = input(
        "Please enter the column (of the source document) from which the information about the boxes begin: ")
   
    boxInformationOrder = input(
        "Please enter the order of the information presented in the source document (Example: L,H,B,W): ")

    configFileName = input(
        "Please enter the name of the config file to be generated: ")

    # Resolving the Paths
    # TODO: Test if these are working
    currentDir = Path('.')
    sourceFile = currentDir / 'Spreadsheets' / sourceFile
    configFile = currentDir / 'Configs' / (configFileName + '.toml')

    workBook = openpyxl.load_workbook(sourceFile, read_only=True)
    workSheet = workBook[sourceSheet]

    workRows = workSheet.max_row
    workColumns = workSheet.max_column

    boxStartCol = openpyxl.utils.column_index_from_string(sourceBoxesStartFrom)
    boxEndCol = boxStartCol + (4 * numberOfBoxes)

    # create the toml file
    file = open(configFile, "w+")

    writeSheetOptions(file, sourceSheet, workRows, workColumns)
    writeMaps(file, workSheet, workColumns, boxStartCol, boxEndCol, boxInformationOrder)
    writeHeaders(file, numberOfBoxes)

    print("Config file created successfully!")

    file.close()


if __name__ == "__main__":
    main()
