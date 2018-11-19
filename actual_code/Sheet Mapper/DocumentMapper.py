import toml
import openpyxl

from pathlib import Path

# package imports
from mapHeaders import mapHeaders


def main():
    # user input
    configDir = input('Enter config file name: ') + ".toml"
    fileDir = input('Enter source file name: ') + ".xlsx"
    outPutFileName = input(
        'Enter the name of the output file to be generated: ') + ".xlsx"

    # TODO: Implement this - currently not working
    # outPutDir = input(
    #     'Enter the name of the subdirectory you want the file to be saved on (Press ENTER for default): ')
    outputSheetName = input('Enter the name of the sheet to be generated: ')

    # Resolve Paths
    currentDir = Path('./..')
    configDir = currentDir / 'Configs' / configDir
    fileDir = currentDir / 'Spreadsheets' / fileDir
    # outPutDir = currentDir / 'Generated' / \
    #     outPutDir.mkdir(parents=True, exist_ok=True)
    #outPutDir = currentDir / 'Generated' /  # outPutDir
    # outPutFileName = outPutDir / outPutFileName
    outPutFileName = currentDir / 'Generated' / outPutFileName # FIXME: fix this to also work with custom directories

    # translate config to dictionary
    configDict = toml.load(configDir)

    # declaring source information variables
    sourceConfig = configDict.get('sheet')
    dataStartRow = 0
    dataEndRow = 0
    skipRows = []
    sourceCol = 0
    sourceSheet = 'Sheet 1'  # default Sheet name in Excel

    # getting the worksheet to transform
    sourceWorkbook = openpyxl.load_workbook(fileDir, data_only=True)

    # loading up the values of the source sheet from the config
    for key, value in sourceConfig.items():
        key = str(key)
        value = str(value)

        if (key == 'datastartrow'):
            dataStartRow = int(value)
        elif (key == 'dataendrow'):
            dataEndRow = int(value)
        elif (key == 'skiprows'):
            if (value != ''):
                skipRows = [int(x) for x in value.split(',')]
        # END OF NEW THINGS

        elif (key == 'cols'):
            sourceCol = openpyxl.utils.column_index_from_string(value)
        elif (key == 'name'):
            sourceSheet = sourceWorkbook[value]

    # output excel destination
    wb = openpyxl.Workbook()
    outputSheet = mapHeaders(
        configDict, wb.create_sheet(title=outputSheetName))

    mapsTo = configDict.get('mapsto')

    # This is where you will start writing on the ouput sheet
    currentWriteRow = 3

    # TODO: Implemented a Feature that ignores Empty Rows. Check if it is working as intended

    # iterating through the source sheet
    for iterRow in range(dataStartRow, dataEndRow + 1):
        outputDict = {}  # empty dictionary for every row

        if iterRow in skipRows:
            continue

        isRowEmpty = True
        for iterCol in range(1, sourceCol + 1):
            colLetter = openpyxl.utils.get_column_letter(iterCol)

            # TODO: check if this works, because you changed read mode to Data only and changed this from internal_value to value

            content = sourceSheet.cell(row=iterRow, column=iterCol).value

            if (content == None):
                content = ''
            elif (isRowEmpty == True):
                isRowEmpty = False

            # Check if the present column maps to something in the config file
            # This is just a sanity check. In reality, the config generator will always have all of the source columns
            if colLetter in mapsTo:
                outputCol = mapsTo[colLetter]  # getting the output columns
                if (outputCol == ""):
                    continue

                # Put the content of the source for every column the source maps to
                for singleColumn in str(outputCol).split(','):
                    outputDict[singleColumn] = [content]

        for key, dictValue in outputDict.items():
            # adjusting to fit the headers
            columnIndex = openpyxl.utils.column_index_from_string(key)
            formattedValue = str(dictValue[0])

            # TODO: Test this, because we implemented a write row inpedendent of the iterRow

            # inserting cells into the new
            outputSheet.cell(row=currentWriteRow,
                             column=columnIndex, value=formattedValue)

        if (isRowEmpty == False):
            currentWriteRow += 1

    wb.save(outPutFileName)
    print("Mapped Document has been generated!")


if __name__ == "__main__":
    main()
