import toml
import openpyxl

from pathlib import Path

# package imports
from mapHeaders import mapHeaders

def main():
    # user input
    configDir = input('Enter config file name: ')
    fileDir = input('Enter source file name: ')
    outputDir = input('Enter the name of the output file to be generated: ')
    outputSheetName = input('Enter the name of the sheet to be generated: ')

    # Resolve Paths
    currentDir = Path('./..')
    configDir = currentDir / 'Configs' / configDir
    fileDir = currentDir / 'Spreadsheets' / fileDir
    outputDir = currentDir / 'Generated' / outputDir

    # declaring source information variables
    sourceConfig = configDict.get('sheet')
    sourceRow = 0
    sourceCol = 0
    sourceSheet = 'Sheet 1' # default Sheet name in Excel

    # getting the worksheet to transform
    sourceWorkbook = openpyxl.load_workbook(fileDir, data_only=True)

    # loading up the values of the source sheet from the config
    for key, value in sourceConfig.items():
        key = str(key)
        value = str(value)

        if (key == 'rows'):
            sourceRow = value
        elif (key == 'cols'):
            sourceCol = openpyxl.utils.column_index_from_string(value)
        elif (key == 'name'):
            sourceSheet = sourceWorkbook.get_sheet_by_name(value)

    # translate config to dictionary
    configDict = toml.load()

    # output excel destination
    wb = openpyxl.Workbook()
    outputSheet = mapHeaders(configDict, wb.create_sheet(title=outputSheetName))

    mapsTo = configDict.get('mapto')

    # iterating through the source sheet
    for iterRow in range(1, sourceRow + 1):
        outputDict = {} # empty dictionary for every row
  
        for iterCol in range(1, sourceCol + 1):
            colLetter = openpyxl.utils.get_column_letter(iterCol)

            # TODO: check if this works, because you changed read mode to Data only and changed this from internal_value to value
          content = sourceSheet.cell(row = iterRow, column = iterCol).value
        
          # Check if the present column maps to something in the config file
          # This is just a sanity check. In reality, the config generator will 
          # always have all of the source columns
          if colLetter in mapsTo:
              outputCol = mapsTo[colLetter] # getting the output columns

              # Put the content of the source for every column the source maps to
              for singleColumn in str(outputCol).split(','):
                  outputDict[singleColumn] = [content]

        for key, dictValue in outputDict.items():
            # adjusting to fit the headers
            normalizedRow = iterRow + 2
            columnIndex = openpyxl.utils.column_index_from_string(key)
            formattedValue = str(dictValue[0])
    
            # inserting cells into the new
            outputSheet.cell(row = normalizedRow, column = columnIndex, value = formattedValue)

    wb.save(outputDir)


if __name__ == "__main__":
    main()
