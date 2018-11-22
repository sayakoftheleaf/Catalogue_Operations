from pathlib import Path
import openpyxl as pyx
from copy import deepcopy


def findLastRowWithMeaningfulValue(inputSheet):
    # Run through the rows
    for inputRow in range(inputSheet.max_row, 0, -1):
        # Run through the columns
        for inputColumn in range(inputSheet.max_column, 0, -1):
            content = inputSheet.cell(row=inputRow, column=inputColumn).value

            # Return inputRow at the first instance when content is found
            if (str(content) != 'None' and str(content) != ''):
                return inputRow

    return 0  # This happens when the sheet is empty


def readAndOutputColumn(
        inputColumn, inputSheet, outputSheet, maxRow, nextWriteRow, writeParameters, headerDict):

    # global emptyRowsInCurrentSheet
    # global nonEmptyRowsInCurrentSheet

    # assuming that headers are in the first row of the input sheet
    columnHeader = inputSheet.cell(row=1, column=inputColumn).value

    # add a new header to the dictionary
    if not(columnHeader in headerDict):
        headerDict[columnHeader] = writeParameters['nextWriteColumn']
        outputSheet.cell(
            row=1, column=writeParameters['nextWriteColumn'], value=columnHeader)
        writeParameters['nextWriteColumn'] += 1

    outputColumn = headerDict[columnHeader]

    # write all the contents of that column in the corresponding output sheet
    for currentRow in range(2, maxRow + 1):
        content = inputSheet.cell(row=currentRow, column=inputColumn).value

        # Leave cells blank instead of printing None
        if (str(content) == 'None'):
            content = ''

        outputSheet.cell(row=nextWriteRow,
                         column=outputColumn, value=str(content))

        nextWriteRow += 1


def mergeOneSheet(inputSheet, outputSheet, writeParameters, headerDict):
    nextWriteRow = deepcopy(writeParameters['nextWriteRow'])

    # TODO: Handle cases when the sheet is empty
    lastRow = findLastRowWithMeaningfulValue(inputSheet)

    for currentColumn in range(1, inputSheet.max_column + 1):

        # Refresh the write row when an entire column is done
        nextWriteRow = writeParameters['nextWriteRow']

        readAndOutputColumn(currentColumn, inputSheet,
                            outputSheet, lastRow, nextWriteRow, writeParameters, headerDict)

    # the next sheet needs to print content below the current sheet
    writeParameters['nextWriteRow'] = nextWriteRow + 1

    # for emptyRow in emptyRowsInCurrentSheet:
    #     outputSheet.delete_rows(emptyRow)

 
    # emptyRowsInCurrentSheet = list(set(emptyRowsInCurrentSheet))

    # - len(emptyRowsInCurrentSheet)
    # print('maxRow is {0} and emptyRowsAre {1} '.format(
    # inputSheet.max_row, emptyRowsInCurrentSheet))

    # emptyRowsInCurrentSheet = []
    # nonEmptyRowsInCurrentSheet = []


def mergeSheets(currentDir, stateObject, outputSheet):
    # Standard Excel worksheet format
    # start writing from row 3, row 1 being for the header and row 2 blank
    writeParameters = {
        'nextWriteRow': 3,
        'nextWriteColumn': 1
    }

    headerDict = {}

    # For every file, run through their sheets
    for inputFile, inputSheets in stateObject['fileAndSheetDict'].items():

        fileDir = currentDir / 'Spreadsheets' / inputFile
        sourceWorkbook = pyx.load_workbook(fileDir, data_only=True)

        # For every sheet to merge
        for sheet in inputSheets.split(','):
            if sheet in stateObject['dontMerge']:
                continue

            # Putting the contents of the current sheet into the output sheet
            mergeOneSheet(sourceWorkbook[sheet],
                          outputSheet, writeParameters, headerDict)
