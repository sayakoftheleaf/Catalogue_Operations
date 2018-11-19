import openpyxl
from pathlib import Path
from os import listdir

# new dictionary that has headers as keys and the column they output
# to in the output file as their values
fileAndSheetDict = {}
headerDict = {}

# Global variables that keep track of the write pointer in the output sheet
nextWriteRow = 3
nextWriteColumn = 1


def correctExtenstion(someString):
    if (len(someString.split('.')) != 2 or someString.split[1] != 'xlsx'):
        return someString.split('.')[0] + '.xlsx'
    else:
        return someString


def addFileToDict(fileName, sheetNames):
    global fileAndSheetDict

    fileAndSheetDict[fileName] = sheetNames


def acceptMultipleFiles():
    numberOfFilesToMerge = input(
        'Enter the number of excel files you need to merge: ')

    for i in range(0, int(numberOfFilesToMerge)):
        fileNumber = str(i + 1)
        fileName = input(
            'Enter the name of file {0} :'.format(fileNumber))
        inputSheets = input(
            'Enter the names of the sheets in file {0}'.format(fileNumber)
            + 'to be merged (multiple files can be separated by commas): ')

        # removing whitespace
        inputSheets = inputSheets.replace(', ', ',')

        # correcting the input Name
        fileName = correctExtenstion(fileName)

        addFileToDict(fileName, inputSheets)


def acceptFilesFromDirectory(currentDir):
    dirName = input('Please enter the name of the directory that has the files to be merged: ')
    sheetVerification = input('Are all the sheets in the files to be merged? (y for YES or n for NO): ')
    if (sheetVerification == 'n' or sheetVerification == 'N'):
        print('Please delete the unnecessary sheets and try again.')
        exit() # Currently cannot handle unnecessary sheets
    
    filesToOpen = listdir(currentDir / 'SpreadSheets' / dirName)
    
    for currentFile in filesToOpen:
        currentWorkbook = openpyxl.load_workbook(currentFile)
        sheets = ','.join(currentWorkbook.get_sheet_names())
        addFileToDict(currentFile, sheets)


def acceptInputAndFormFileDict(curentDir):
    navigationType = input(
        'Are you going to merge individual files or all files in a directory (f for file, d for dir): ')

    if (navigationType == 'd'):
        acceptFilesFromDirectory(curentDir)
    elif (navigationType == 'f'):
        acceptMultipleFiles()
    else:
        print('Invalid Input. Please try again.')
        acceptInputAndFormFileDict(curentDir)


# Writes all the contents of a column into the output Sheet


def readAndOutputColumn(inputColumn, inputSheet, outputSheet, maxRow):
    global headerDict
    global nextWriteColumn
    global nextWriteRow

    # assuming that headers are in the first row of the input sheet
    columnHeader = inputSheet.cell(row=1, column=inputColumn).value

    # add a new header to the dictionary
    if not(columnHeader in headerDict):
        headerDict[columnHeader] = nextWriteColumn
        outputSheet.cell(row=1, column=nextWriteColumn, value=columnHeader)
        nextWriteColumn += 1

    outputColumn = headerDict[columnHeader]

    isRowEmpty = True  # becomes False when first content is encountered

    # write all the contents of that column in the corresponding output sheet
    for currentRow in range(2, maxRow + 1):
        content = inputSheet.cell(row=currentRow, column=inputColumn).value

        # Leave cells blank instead of printing None
        # None clutters up the sheet
        if (content == 'None'):
            content = ''
        elif (isRowEmpty == True):
            isRowEmpty = False

        outputSheet.cell(row=nextWriteRow,
                         column=outputColumn, value=str(content))

        # if the Row is empty, you can overwrite it next time
        if (isRowEmpty == False):
            nextWriteRow += 1


def mergeSheet(inputSheet, outputSheet):
    global nextWriteColumn
    global nextWriteRow

    startWritingAtRow = nextWriteRow

    # for every column in the current sheet
    for currentColumn in range(1, inputSheet.max_column + 1):
        # Since this is writing one column at a time, you need to refresh
        # the write row after a column is done
        nextWriteRow = startWritingAtRow

        readAndOutputColumn(currentColumn, inputSheet,
                            outputSheet, inputSheet.max_row)

    # the next sheet needs to print content below the current sheet
    nextWriteRow += inputSheet.max_row + 1


def main():

    currentDir = Path('./..')

    # accept the Input Files and Sheets
    acceptInputAndFormFileDict(currentDir)

    outputFile = input('Enter the name of the output file to be generated: ')
    outputSheet = input('Enter the name of the output sheet to be generated: ')

    # auto-correcting the output extension and resolving the Path
    outputFile = currentDir / 'Generated' / correctExtenstion(outputFile)

    outputWorkbook = openpyxl.Workbook()
    writeSheet = outputWorkbook.create_sheet(title=outputSheet)

    # For every file, run through their sheets
    for inputFile, inputSheets in fileAndSheetDict:

        fileDir = currentDir / 'Spreadsheets' / inputFile
        sourceWorkbook = openpyxl.load_workbook(fileDir, data_only=True)

        # For every sheet to merge
        for sheet in inputSheets.split(','):
            # TODO: There has to be a better way to do this
            # Putting the contents of the current sheet into the output sheet
            inputSheet = sourceWorkbook[sheet]
            mergeSheet(inputSheet, writeSheet)

    outputWorkbook.save(outputFile)
    print('Sheets have been merged and output file has been generated!')


if __name__ == "__main__":
    main()
