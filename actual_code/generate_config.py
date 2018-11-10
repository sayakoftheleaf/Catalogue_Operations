import toml
import openpyxl


def writeSheetOptions(fileObject, name, rows, cols):
    commentString = "\n# Details of the source sheet. Edit in case info is wrong \n\n"
    name = "sheet.name = {0}\n".format(name)
    rows = "sheet.rows = +{0}\n".format(str(rows))
    cols = "sheet.cols = {0}\n".format(openpyxl.utils.get_column_letter(cols))

    fileObject.write("{0}{1}{2}{3}".format(commentString, name, rows, cols))


# TODO: Implement cases where one file maps to multiple sources


def writeHeaders(fileObject, sheet, lastCol):
    colString = "\n# Put all the mappings here\n If one column maps to several columns, separate them with a comma\n DO NOT PUT WHITESPACE AFTER COMMA!\n For example: mapsto.A = \"B,X\"\n\n"
    # iterate through all of the columns
    for currentCol in range(1, lastCol + 1):
        # get the letter equivalent of the current column
        columnLetter = openpyxl.utils.get_column_letter(currentCol)
        # get the header inside the cell
        header = sheet.cell(row=1, column=currentCol).value
        colString += "mapsto.{0} = \"\" # {1}\n".format(columnLetter, header)

    fileObject.write(colString)


def writeMaps(fileObject):


def main():
    sourceFile = input("Please enter the name of the source Excel document: ")
    sourceSheet = input(
        "Please enter the name of the sheet you want to be mapped: ")
    # numberOfBoxes = input(
    # "Please enter the number of boxes that need to be shipped: ")
    configFileName = input(
        "Please enter the name of the config file to be generated: ")

    workBook = openpyxl.load_workbook(sourceFile, read_only=True)
    workSheet = workBook[sourceSheet]

    workRows = workSheet.max_row
    workColumns = workSheet.max_column

    # create the toml file
    fileName = configFileName + ".toml"
    file = open(fileName, "w+")

    writeSheetOptions(file, sourceSheet, workRows, workColumns)
    writeHeaders(file, workSheet, workColumns)

    print("Config file created successfully!")

    file.close()


if __name__ == "__main__":
    main()
