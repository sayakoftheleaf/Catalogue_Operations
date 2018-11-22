import toml
import openpyxl

from generate_box_config import generateExcelColumns

def writeMaps(fileObject, sheet, lastCol, headerRow, boxStartCol, boxEndCol,  boxInformationOrder):

    commentString = "\n# Put all the mappings here\n# If one column maps to several columns, separate them with a comma\n# DO NOT PUT WHITESPACE AFTER COMMA!\n# For example: mapsto.A = \"B,X\"\n\n"

    # intitializing the final String with the beginning comment
    colString = commentString

    orderDict = evaluateOrder(boxInformationOrder)

    currentCol = 1
    outputBoxCol = "T"

    # iterate through all of the columns
    while (currentCol <= lastCol):

        if (currentCol >= boxStartCol and currentCol < boxEndCol):

            # printing comment to indicate beginning of box mappings
            if (currentCol == boxStartCol):
                colString += "\n# beginning of the box mappings. Please do not change these.\n"

            columnLetter = [openpyxl.utils.get_column_letter(currentCol)]
            columnLetter.append(openpyxl.utils.get_column_letter(currentCol+1))
            columnLetter.append(openpyxl.utils.get_column_letter(currentCol+2))
            columnLetter.append(openpyxl.utils.get_column_letter(currentCol+3))

            BoxCols = [generateExcelColumns(outputBoxCol)]
            BoxCols.append(generateExcelColumns(BoxCols[0]))
            BoxCols.append(generateExcelColumns(BoxCols[1]))
            BoxCols.append(generateExcelColumns(BoxCols[2]))

            boxIter = 0
            for key, value in orderDict.items():
              colString += "mapsto.{0} = \"{1}\" \n".format(columnLetter[value], BoxCols[boxIter])
              boxIter += 1

            # increments
            currentCol += 4
            outputBoxCol = BoxCols[3]

            # printing comment to indicate the end of box mappings
            if (currentCol > boxEndCol):
                colString += "# end of the box mappings.\n\n"
        else:
            # get the header inside the cell
            header = sheet.cell(row=int(headerRow), column=currentCol).value

            if (header != None):
                header = header.replace('\n', ' ')
                columnLetter = openpyxl.utils.get_column_letter(currentCol)
                colString += "mapsto.{0} = \"\" # {1}\n".format(
                    columnLetter, header)

            # increment
            currentCol += 1

    fileObject.write(colString)


def evaluateOrder(boxInformationOrder):
    # Removing commas and spaces and converting everything to uppercase
    boxInformationOrder = boxInformationOrder.replace(" ", "")
    boxInformationOrder = boxInformationOrder.replace(",", "")
    boxInformationOrder = boxInformationOrder.upper()

    orderDict = {}
    for value in ['L', 'B', 'H', 'W']:
        orderDict[value] = boxInformationOrder.index(value)

    return orderDict
