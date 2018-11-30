from pathlib import Path
import openpyxl as pyx
from copy import deepcopy

from SourceChecks import checkForRepeatColumns

def findLastRowWithMeaningfulValue(inputSheet):
    # Run through the rows
    for inputRow in range(inputSheet.max_row, 0, -1):
        # Run through the columns
        for inputColumn in range(inputSheet.max_column, 0, -1):
            content = inputSheet.cell(row=inputRow, column=inputColumn).value

            # Return inputRow at the first instance when content is found
            if (str(content) != 'None' and str(content) != ''):
                return inputRow

    return 0  # This happens when the sheet is empty


def makeNewHeadersAndMapColumns(
    inputSheet, outputSheet, writeParameters, headerDict, duplicates, columnMappings, debugMode):

    # Add an extra column that puts the source of the contents for better
    # debugging
    if (debugMode == 'y' and headerDict.get('SourceRow') == None):
        headerDict['SourceRow'] = writeParameters['nextWriteColumn']
        outputSheet.cell(
                row=1, column=writeParameters['nextWriteColumn'], value="source of this row")
        writeParameters['nextWriteColumn'] += 1

    for col in range(1, inputSheet.max_column + 1):
        columnHeader = inputSheet.cell(row=1, column=col).value
        if not(columnHeader == None):
             columnHeader = columnHeader.strip() # trim whitespace
             columnHeaderLower = columnHeader.lower()
            
        # if this column is a duplicate
        if columnHeaderLower in duplicates:
            # the first encounter with a duplicate is the same as that of a
            # non duplicate
            if (duplicates[columnHeaderLower] == 0):
                duplicates[columnHeaderLower] = 1
            # modify the header for all future encounters
            else:
                duplicateExtension = '[{0}]'.format(str(duplicates[columnHeaderLower]))
                
                # next time, use the next duplicate index
                duplicates[columnHeaderLower] += 1

                columnHeader = columnHeader + duplicateExtension
                columnHeaderLower = columnHeaderLower + duplicateExtension

        # add a new header to the dictionary
        if not(columnHeaderLower in headerDict):
            headerDict[columnHeaderLower] = writeParameters['nextWriteColumn']
            outputSheet.cell(
                row=1, column=writeParameters['nextWriteColumn'], value=columnHeader)
            columnMappings[col] = writeParameters['nextWriteColumn']
            writeParameters['nextWriteColumn'] += 1
        else:       
            columnMappings[col] = headerDict[columnHeaderLower]

def mergeOneSheet(
    inputSheet, outputSheet, writeParameters, headerDict, duplicates, debugMode, fileName, sheetName):
    columnMappings = {}

    # TODO: Handle cases when the sheet is empty
    lastRow = findLastRowWithMeaningfulValue(inputSheet)

    makeNewHeadersAndMapColumns(
            inputSheet, outputSheet, writeParameters, headerDict, duplicates, columnMappings, debugMode)

    for row in range(2, lastRow + 1):
        isRowEmpty = True
        
        if (debugMode == 'y'):
            outputSheet.cell(
                row = writeParameters['nextWriteRow'],
                column = headerDict['SourceRow'],
                value = fileName + '=>' + sheetName + ': row = ' + str(row))
        for col in range(1, inputSheet.max_column + 1):
            content = inputSheet.cell(row=row, column=col).value
            if (str(content) != 'None' and str(content) != ''):
                isRowEmpty = False
            
            outputColumn = columnMappings[col] 
            outputSheet.cell(row=writeParameters['nextWriteRow'], column=outputColumn, value=content)
        
        if (isRowEmpty == False):
            writeParameters['nextWriteRow'] += 1

def mergeSheets(currentDir, stateObject, outputSheet):
    # Standard Excel worksheet format
    # start writing from row 3, row 1 being for the header and row 2 blank
    writeParameters = {
        'nextWriteRow': 3,
        'nextWriteColumn': 1
    }

    headerDict = {}

    # For every file, run through their sheets
    for inputFile, inputSheets in stateObject['fileAndSheetDict'].items():

        fileDir = currentDir / 'Spreadsheets' / inputFile
        sourceWorkbook = pyx.load_workbook(fileDir, data_only=True)

        # For every sheet to merge
        for sheet in inputSheets.split(','):
            duplicates = checkForRepeatColumns(sourceWorkbook[sheet], sheet, inputFile)
            # Putting the contents of the current sheet into the output sheet
            mergeOneSheet(
                sourceWorkbook[sheet], outputSheet, writeParameters, headerDict, duplicates,
                stateObject['debugMode'], inputFile, sheet)
