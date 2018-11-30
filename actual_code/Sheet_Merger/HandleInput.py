from os import listdir
import openpyxl as pyx

# A bunch of verification checks prompted to the user to ensure correct
# input file format
def fileChecks():
    xlsxVerification = input(
        'Are all the files in .xlsx format? (.xls is not supported) (y or n): ')

    if (xlsxVerification == 'n' or xlsxVerification == 'N'):
        print('Please convert the files to .xlsx and try again.')
        exit()

    headerVerification = input(
        'Do all the sheets have headers ONLY on Row 1 and no other row? (y or n): ')

    if (headerVerification == 'n' or headerVerification == 'N'):
        print('Please modify the headers and put them on the first row before running this program.')
        exit()


def mergeAllSheets(currentFile):
    allSheets = input(
        'all are the sheets in the file' + currentFile + 'to be merged? (y or n)')
    if (allSheets.lower() == 'y'):
        return True
    elif (allSheets.lower() == 'n'):
        return False
    else:
        print('invalid input')
        sheetInputs(currentFile)

# Add an extension to the filename entered as input if it already doesn't have one
# Or correct a wrong extension hoping it was a mistype


def correctExtenstion(someString):
    if (len(someString.split('.')) != 2 or someString.split[1] != 'xlsx'):
        return someString.split('.')[0] + '.xlsx'
    else:
        return someString

# Create a dictionary of the files and the sheets in them that need to be merged


def addFileToDict(fileName, sheetNames, fileAndSheetDict):
    fileAndSheetDict[fileName] = sheetNames


def acceptFilesFromDirectory(currentDir, fileAndSheetDict):

    dirName = input(
        'Please enter the name of the directory that has the files to be merged: ')

    fileChecks()

    for currentFile in listdir(currentDir / 'Spreadsheets' / dirName):
        extension = currentFile.split('.')
        extension = extension[len(extension) - 1]

        if (extension == 'xlsx'):
            fullPath = currentDir / 'Spreadsheets' / dirName / currentFile
            currentWorkbook = pyx.load_workbook(fullPath)

            sheetNames = currentWorkbook.sheetnames

            if (mergeAllSheets(currentFile) == False):
                skipThese = input(
                    'enter sheet names in the file ' +
                    currentFile + ' that are to be skipped (separate with commas)')

                # remove the sheets to skip from the sheet list
                for skipThis in skipThese.split(','):
                    skipThis = skipThis.strip()  # trim string
                    if skipThis in sheetNames:
                        sheetNames.remove(skipThis)
                    else:
                        print(
                            'sheet ' + skipThis + ' not found in file ' + currentFile + '. Continuing.')

            # make a string out of the sheet list
            sheets = ','.join(sheetNames)

            addFileToDict(dirName + '/' + currentFile,
                          sheets, fileAndSheetDict)


def acceptMultipleFiles(currentDir, fileAndSheetDict):
    fileChecks()
    dirPath = currentDir / 'SpreadSheets'

    numberOfFilesToMerge = input(
        'Enter the number of excel files you need to merge: ')

    for i in range(0, int(numberOfFilesToMerge)):
        fileNumber = str(i + 1)
        fileName = input(
            'Enter the name of file {0} :'.format(fileNumber))
        subDirPath = input(
            'Enter the subdirectory path where the file is (leave blank for the main directory)') or False

        if (subDirPath == False):
            dirPath = dirPath / fileName
        else:
            dirPath = dirPath / subDirPath / fileName

        # correcting the input Name
        fileName = correctExtenstion(fileName)

        if (mergeAllSheets(fileName) == False):
            inputSheets = input(
                'Enter the names of the sheets in file {0}'.format(fileNumber)
                + 'to be merged (multiple files can be separated by commas): ')
            # removing whitespace
            inputSheets = inputSheets.replace(', ', ',')
        else: 
            currentWorkbook = pyx.load_workbook(dirPath)
            inputSheets = currentWorkbook.sheetnames
            inputSheets = inputSheets.join(',')

        addFileToDict(fileName, inputSheets, fileAndSheetDict)


def acceptInputAndFormFileDict(curentDir, stateObject):

    navigationType = input(
        'Are you going to merge individual files'
        'or all files in a directory (f for file, d for dir): ')

    stateObject['debugMode'] = input(
        'Do you want to map the source row to the output row for debugging?'
        '(y or n):')

    if (navigationType == 'd'):
        acceptFilesFromDirectory(curentDir, stateObject['fileAndSheetDict'])
    elif (navigationType == 'f'):
        acceptMultipleFiles(currentDir, stateObject['fileAndSheetDict'])
    else:
        print('Invalid Input. Please try again.')
        acceptInputAndFormFileDict(curentDir, stateObject)
