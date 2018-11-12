import toml
import openpyxl

from generate_box_config import generateExcelColumns


def writeSheetOptions(fileObject, name, rows, cols):
    commentString = "\n# Details of the source sheet. Edit in case info is wrong \n\n"
    name = "sheet.name = \"{0}\"\n".format(name)
    rows = "sheet.rows = +{0}\n".format(str(rows))
    cols = "sheet.cols = \"{0}\"\n".format(
        openpyxl.utils.get_column_letter(cols))

    fileObject.write("{0}{1}{2}{3}".format(commentString, name, rows, cols))


# TODO: Implement cases where one file maps to multiple sources

def writeMaps(fileObject, sheet, lastCol):
    commentString = "\n# Put all the mappings here\n# If one column maps to several columns, separate them with a comma\n# DO NOT PUT WHITESPACE AFTER COMMA!\n# For example: mapsto.A = \"B,X\"\n\n"
    colString = commentString
    # iterate through all of the columns
    for currentCol in range(1, lastCol + 1):
        # get the letter equivalent of the current column
        columnLetter = openpyxl.utils.get_column_letter(currentCol)
        # get the header inside the cell
        header = sheet.cell(row=1, column=currentCol).value
        colString += "mapsto.{0} = \"\" # {1}\n".format(columnLetter, header)

    fileObject.write(colString)


def evaluateOrder(boxInformationOrder):
    # Removing commas and spaces and converting everything to uppercase
    boxInformationOrder = boxInformationOrder.replace(" ", "")
    boxInformationOrder = boxInformationOrder.replace(",", "")
    boxInformationOrder = boxInformationOrder.upper()

    orderDict = {}
    for value in ['L', 'B', 'H', 'W']:
        orderDict[value] = boxInformationOrder.index(value)

    return orderDict

def mapBoxes (fileObject, sheet, sourceBoxesStartFrom, boxInformationOrder):
    
    


def writeHeaders(fileObject, numberOfBoxes):
    commentString = "\n# These are the headers for the output file. Please do not change these.\n"
    returnedDict = headersOfBoxes(numberOfBoxes, headersBeforeBox([]))
    mapsArr = headersAfterBoxes(
        returnedDict['currentColumn'], returnedDict['mapsto'])

    mapsString = commentString
    for line in mapsArr:
        mapsString += "header." + line + "\n"

    fileObject.write(mapsString)


def headersBeforeBox(mapsArr):
    mapsArr.append("A = \"Supplier Status (Mandatory)\"")
    mapsArr.append("B = \"UPC/EAN (Mandatory)\"")
    mapsArr.append("C = \"Identifier Type (Mandatory)\" ")
    mapsArr.append("D = \"Override GS1 Check?\" ")
    mapsArr.append("E = \"MPN (Mandatory)\" ")
    mapsArr.append(
        "F = \"Product Name (Original as Per Supplier) (Mandatory)\"")
    mapsArr.append("G = \"Brand (Mandatory)\" ")
    mapsArr.append("H = \"Product Category\" ")
    mapsArr.append("I = \"Dropship Gross Purchase Price\" ")
    mapsArr.append("J = \"Dropship Discount\" ")
    mapsArr.append("K = \"Dropship Net Purchase Price\"")
    mapsArr.append("L = \"Dropship Supplier Shipping Charge SKU Flat Rate\"")
    mapsArr.append("M = \"Dropship Supplier Shipping Charge SKU % Rate\"")
    mapsArr.append("N = \"Dropship Supplier Handling Charge Rate\"")
    mapsArr.append("O = \"Dropship Outbound Shipment Type\"")
    mapsArr.append("P = \"Fragile\"")
    mapsArr.append("Q = \"Override Shipping Method?\"")
    mapsArr.append("R = \"Override Dropship Outbound Carrier\"")
    mapsArr.append("S = \"Override Dropship Order Handling Time\"")
    mapsArr.append("T = \"# of Cartons\"")

    return mapsArr


def headersOfBoxes(numberOfBoxes, mapsArr):
     # starting with the column before the actual start for convenience
    currentColumn = "T"
    for iter in range(1, numberOfBoxes + 1):
        currentColumn = generateExcelColumns(currentColumn)
        mapsArr.append(currentColumn + " = \"Box " + str(iter) +
                       " Shipping Length (in.) (Mandatory)\"")

        currentColumn = generateExcelColumns(currentColumn)
        mapsArr.append(currentColumn + " = \"Box " + str(iter) +
                       " Shipping Breadth (in.) (Mandatory)\"")

        currentColumn = generateExcelColumns(currentColumn)
        mapsArr.append(currentColumn + " = \"Box " + str(iter) +
                       " Shipping Height (in.) (Mandatory)\"")

        currentColumn = generateExcelColumns(currentColumn)
        mapsArr.append(currentColumn + " = \"Box " + str(iter) +
                       " Shipping Weight (lbs.) (Mandatory)\"")

    return {
        'mapsto': mapsArr,
        'currentColumn': currentColumn
    }


def headersAfterBoxes(currentColumn, mapsArr):
    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Bulk Gross Purchase Price\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Bulk Discount\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Bulk Net Purchase Price\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Case Pack/Master Carton (Units)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn +
                   " = \"Case Pack / Master Carton Shipping L (in.)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn +
                   " = \"Case Pack / Master Carton Shipping B (in.)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn +
                   " = \"Case Pack / Master Carton Shipping H (in.)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn +
                   " = \"Case Pack / Master Carton Shipping W (lbs.)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Available in Warehouse #1\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Available in Warehouse #2\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Available in Warehouse #3\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"MAP\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Harmonized Code\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Listing attributes\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"MOQ\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Tariff #\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Freight Class\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Keywords\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Color/Finish\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Country of Manufacture\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"MSRP\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Primary Material\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Distressed Finish? (y/n)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Assembly Required? (y/n)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Assembled Item Dimensions (Width)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Assembled Item Dimensions (Depth)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Assembled Item Dimensions (Height)\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Assembled Item Weight\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Item Description\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 1\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 2\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 3\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 4\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 5\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 6\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 7\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 8\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"feature 9\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 1\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 2\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 3\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 4\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 5\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Image 6\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Collection\"")

    currentColumn = generateExcelColumns(currentColumn)
    mapsArr.append(currentColumn + " = \"Includes\"")

    return mapsArr


def main():
    sourceFile = input("Please enter the name of the source Excel document: ")
    sourceSheet = input(
        "Please enter the name of the sheet you want to be mapped: ")
    numberOfBoxes = input(
        "Please enter the number of boxes that need to be shipped: ")
    sourceBoxesStartFrom = input(
        "Please enter the column (of the source document) from which the information about the boxes begin:")
    boxInformationOrder = input(
        "Please enter the order of the information presented in the source document (Example: L,H,B,W)")

    numberOfBoxes = int(numberOfBoxes)

    configFileName = input(
        "Please enter the name of the config file to be generated: ")

    workBook = openpyxl.load_workbook(sourceFile, read_only=True)
    workSheet = workBook[sourceSheet]

    workRows = workSheet.max_row
    workColumns = workSheet.max_column

    # create the toml file
    fileName = configFileName + ".toml"
    file = open(fileName, "w+")

    writeSheetOptions(file, sourceSheet, workRows, workColumns)
    writeMaps(file, workSheet, workColumns)
    writeHeaders(file, numberOfBoxes)

    print("Config file created successfully!")

    file.close()


if __name__ == "__main__":
    main()
